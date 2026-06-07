"""Unit-Tests fuer den Excel-Exporter (save_campaign_as_excel).
Laufen vollstaendig offline ohne Google-Cloud-Credentials.
"""
import json
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from app.agent import save_campaign_as_excel

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "exports")

SAMPLE = {
    "domain": "test.de",
    "datum": "Januar 2025",
    "kampagnen": [
        {
            "name": "Brand Search",
            "typ": "Suchnetzwerk",
            "budget": "10 Euro/Tag",
            "gebotsstrategie": "Klicks maximieren",
            "begruendung": "Markenschutz.",
        },
        {
            "name": "Generische Suche",
            "typ": "Suchnetzwerk",
            "budget": "30 Euro/Tag",
            "gebotsstrategie": "Conversions maximieren",
            "begruendung": "Neukundengewinnung.",
        },
        {
            "name": "Performance Max (PMax)",
            "typ": "PMax",
            "budget": "20 Euro/Tag",
            "gebotsstrategie": "Conversions maximieren",
            "begruendung": "Skalierung.",
        },
        {
            "name": "Remarketing",
            "typ": "Display",
            "budget": "5 Euro/Tag",
            "gebotsstrategie": "Conversions maximieren",
            "begruendung": "Rueckgewinnung.",
        },
    ],
    "keywords": [
        {
            "kampagne": "Brand Search",
            "anzeigengruppe": "Brand",
            "keyword": "[testbrand]",
            "match_type": "Exact",
            "suchintention": "Navigational",
            "kategorie": "Brand",
            "prioritaet": "Hoch",
        },
        {
            "kampagne": "Brand Search",
            "anzeigengruppe": "Brand",
            "keyword": '"testbrand erfahrungen"',
            "match_type": "Phrase",
            "suchintention": "Navigational",
            "kategorie": "Brand",
            "prioritaet": "Mittel",
        },
        {
            "kampagne": "Generische Suche",
            "anzeigengruppe": "AG 1",
            "keyword": "[produkt kaufen]",
            "match_type": "Exact",
            "suchintention": "Transactional",
            "kategorie": "Produkt",
            "prioritaet": "Hoch",
        },
    ],
    "anzeigentexte": [
        {
            "kampagne": "Brand Search",
            "anzeigengruppe": "Brand",
            "headlines": [
                "Testbrand Offiziell",
                "Jetzt anfragen",
                "Top Service",
                "Faire Preise",
                "Schnelle Lieferung",
            ],
            "descriptions": [
                "Hochwertige Produkte direkt vom Hersteller. Jetzt kostenlos anfragen.",
                "Schnelle Lieferung und fairer Preis. Ueberzeugen Sie sich selbst.",
            ],
        },
        {
            "kampagne": "Generische Suche",
            "anzeigengruppe": "AG 1",
            "headlines": [
                "Produkt jetzt kaufen",
                "Top Angebot sichern",
                "Jetzt bestellen",
            ],
            "descriptions": [
                "Unser Produkt ueberzeugt durch Qualitaet und Preis. Bestellen Sie jetzt.",
            ],
        },
    ],
    "landingpage_analyse": {
        "usps": [
            {"bezeichnung": "USP 1", "beschreibung": "Beschreibung 1"},
            {"bezeichnung": "USP 2", "beschreibung": "Beschreibung 2"}
        ],
        "zielgruppen": [
            {"name": "Zielgruppe 1", "beschreibung": "Zielgruppenbeschreibung 1"}
        ],
        "produkte": [
            {"name": "Produkt 1", "inhalte": "Inhalt 1", "formate": "Format 1", "besonderheiten": "Besonderheit 1"}
        ],
        "tonalitaet": [
            {"merkmal": "Merkmal 1", "auspraegung": "Auspraegung 1"}
        ],
        "keywords": [
            "keyword 1",
            "keyword 2"
        ]
    },
}


@pytest.fixture(scope="module")
def generated_wb():
    filename = "unit_test_exporter.xlsx"
    result = save_campaign_as_excel(filename, json.dumps(SAMPLE))
    assert "exportiert" in result, f"Unerwartetes Ergebnis: {result}"
    filepath = os.path.join(EXPORT_DIR, filename)
    assert os.path.exists(filepath), f"Datei nicht gefunden: {filepath}"
    wb = openpyxl.load_workbook(filepath)
    return wb


def test_sheet_names(generated_wb):
    names = generated_wb.sheetnames
    assert "Landingpage-Analyse" in names
    assert "Kampagnenstruktur" in names
    assert "Keyword-Recherche" in names
    assert "Anzeigentexte (RSA)" in names


def test_landingpage_analyse_sheet(generated_wb):
    ws = generated_wb["Landingpage-Analyse"]

    # Check USPs
    assert ws.cell(row=6, column=1).value == "1"
    assert ws.cell(row=6, column=2).value == "USP 1"
    assert ws.cell(row=6, column=3).value == "Beschreibung 1"
    assert ws.cell(row=7, column=1).value == "2"
    assert ws.cell(row=7, column=2).value == "USP 2"
    assert ws.cell(row=7, column=3).value == "Beschreibung 2"

    # Check Zielgruppen
    assert ws.cell(row=16, column=1).value == "1"
    assert ws.cell(row=16, column=2).value == "Zielgruppe 1"
    assert ws.cell(row=16, column=3).value == "Zielgruppenbeschreibung 1"

    # Check Produkte
    assert ws.cell(row=24, column=1).value == "Produkt 1"
    assert ws.cell(row=24, column=2).value == "Inhalt 1"
    assert ws.cell(row=24, column=3).value == "Format 1"
    assert ws.cell(row=24, column=4).value == "Besonderheit 1"

    # Check Tonalitaet
    assert ws.cell(row=33, column=1).value == "Merkmal 1"
    assert ws.cell(row=33, column=2).value == "Auspraegung 1"

    # Check Keywords
    assert ws.cell(row=39, column=1).value == "keyword 1"
    assert ws.cell(row=40, column=1).value == "keyword 2"

    # Check font formatting (Arial)
    for row in range(6, 44):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                assert cell.font.name == "Arial", f"Zelle ({row},{col}) hat nicht Arial, sondern {cell.font.name}"



def test_gridlines_hidden(generated_wb):
    for name in generated_wb.sheetnames:
        ws = generated_wb[name]
        assert not ws.views.sheetView[0].showGridLines, (
            f"Gitternetzlinien sollten auf Blatt '{name}' ausgeblendet sein"
        )


def test_kampagnenstruktur_headers(generated_wb):
    ws = generated_wb["Kampagnenstruktur"]
    # Zeile 5 = Spaltenkoepfe
    headers = [ws.cell(row=5, column=ci).value for ci in range(1, 6)]
    assert "Kampagne" in headers
    assert "Budget" in headers
    assert "Gebotsstrategie" in headers


def test_kampagnenstruktur_rows(generated_wb):
    ws = generated_wb["Kampagnenstruktur"]
    # Zeile 6 = erste Datenzelle (Brand Search)
    val = ws.cell(row=6, column=1).value
    assert val == "Brand Search"
    # PMax bekommt gruenen Hintergrund E2EFDA
    pmax_row = None
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=1).value and "pmax" in str(ws.cell(row=r, column=1).value).lower():
            pmax_row = r
            break
    if pmax_row:
        fill = ws.cell(row=pmax_row, column=1).fill.fgColor.value
        assert "E2EFDA" in fill.upper() or "e2efda" in fill.lower(), (
            f"PMax-Zeile sollte gruen sein, ist: {fill}"
        )


def test_keyword_sheet_section_headers(generated_wb):
    ws = generated_wb["Keyword-Recherche"]
    # Es sollte mindestens eine verschmolzene Zelle geben (Abschnitts-Header)
    assert len(list(ws.merged_cells.ranges)) > 0


def test_keyword_exact_match_color(generated_wb):
    ws = generated_wb["Keyword-Recherche"]
    # Suche eine Zelle mit Wert "Exact" in Spalte 3 und pruefe Fuellfarbe D9E1F2
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 3 and cell.value == "Exact":
                fill = cell.fill.fgColor.value
                assert "D9E1F2" in fill.upper() or "d9e1f2" in fill.lower(), (
                    f"Exact Match soll D9E1F2 haben, hat: {fill}"
                )
                found = True
    assert found, "Kein 'Exact'-Match-Type-Wert in Spalte 3 gefunden"


def test_keyword_priority_hoch_color(generated_wb):
    ws = generated_wb["Keyword-Recherche"]
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 6 and cell.value == "Hoch":
                fill = cell.fill.fgColor.value
                assert "C6EFCE" in fill.upper() or "c6efce" in fill.lower(), (
                    f"'Hoch'-Prioritaet soll C6EFCE haben, hat: {fill}"
                )
                found = True
    assert found, "Kein 'Hoch'-Prioritaets-Wert in Spalte 6 gefunden"


def test_anzeigentexte_character_count(generated_wb):
    ws = generated_wb["Anzeigentexte (RSA)"]
    # Zeichen-Spalte (col 3) sollte Werte wie "18/30" enthalten
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 3 and cell.value and "/" in str(cell.value):
                found = True
                break
    assert found, "Keine Zeichenanzahl-Zelle (Format X/Y) in Spalte 3 gefunden"


def test_anzeigentexte_status_ok(generated_wb):
    ws = generated_wb["Anzeigentexte (RSA)"]
    # Status-Spalte (col 5) sollte mindestens einmal 'OK' enthalten
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 5 and cell.value == "OK":
                found = True
                break
    assert found, "Kein 'OK'-Status in der Status-Spalte (5) gefunden"


def test_arial_font(generated_wb):
    ws = generated_wb["Kampagnenstruktur"]
    # Datenzeile Zeile 6 soll Arial-Schrift haben
    font_name = ws.cell(row=6, column=1).font.name
    assert font_name == "Arial", f"Schriftart sollte Arial sein, ist: {font_name}"


def test_banner_font_size(generated_wb):
    ws = generated_wb["Kampagnenstruktur"]
    # Zeile 1 = grosser Banner mit font size 15
    font_size = ws.cell(row=1, column=1).font.size
    assert font_size == 15, f"Banner-Schriftgroesse sollte 15 sein, ist: {font_size}"


def test_fallback_old_format():
    """Altes Zeilenformat soll weiterhin funktionieren."""
    old_data = {
        "Strategie": [
            ["Kampagne", "Typ", "Budget"],
            ["Brand", "Suche", "10 Euro"],
        ]
    }
    result = save_campaign_as_excel("unit_test_old_format.xlsx", json.dumps(old_data))
    assert "exportiert" in result
    filepath = os.path.join(EXPORT_DIR, "unit_test_old_format.xlsx")
    wb = openpyxl.load_workbook(filepath)
    assert "Strategie" in wb.sheetnames
