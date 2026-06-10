# ruff: noqa
import datetime
import json
import os
import re
from zoneinfo import ZoneInfo

from google.adk.agents import Agent
from google.adk.models import Gemini
from google.genai import types

from app.common import apply_rules


def save_campaign_as_excel(filename: str, campaign_data_json: str) -> str:
    """Speichert Kampagnendaten als formatierte Excel-Datei (.xlsx) mit Template-Styling.

    Args:
        filename: Name der Excel-Datei (z.B. kampagne_adsforsocial.xlsx)
        campaign_data_json: JSON-String mit semantischen Kampagnendaten. Erwartete Struktur:
            {
              "domain": "adsforsocial.de",
              "datum": "Mai 2025",
              "kampagnen": [
                {"name": "Brand Search", "typ": "Suchnetzwerk", "budget": "15 Euro/Tag",
                 "gebotsstrategie": "Klicks maximieren", "begruendung": "..."}
              ],
              "keywords": [
                {"kampagne": "Brand Search", "anzeigengruppe": "Brand",
                 "keyword": "[adsforsocial]", "match_type": "Exact",
                 "suchintention": "Navigational / Brand", "kategorie": "Brand",
                 "prioritaet": "Hoch"}
              ],
              "anzeigentexte": [
                {"kampagne": "Brand Search", "anzeigengruppe": "Brand",
                 "headlines": ["Headline 1", "Headline 2"],
                 "descriptions": ["Description 1", "Description 2"]}
              ]
            }
    """
    import json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    export_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "exports")
    os.makedirs(export_dir, exist_ok=True)

    if not filename.endswith(".xlsx"):
        filename = re.sub(r"\.[^.]+$", "", filename) + ".xlsx"
    filename = os.path.basename(filename)
    filepath = os.path.join(export_dir, filename)

    try:
        data = json.loads(campaign_data_json)
    except Exception as e:
        return f"Fehler beim Parsen der JSON-Daten: {str(e)}"

    # ---- Farbpalette (aus Template extrahiert) ----
    C_NAVY       = "1F3864"
    C_MEDBLUE    = "2E75B6"
    C_GREEN_DARK = "375623"
    C_GREEN_PMAX = "70AD47"
    C_ORANGE     = "C55A11"
    C_TEAL       = "1F6B75"
    C_YELLOW_DK  = "7F6000"
    C_WHITE      = "FFFFFF"
    C_DARK       = "404040"
    F_GRAY       = "F2F2F2"
    F_LIGHTBLUE  = "DEEAF1"
    F_LIGHTGREEN = "E2EFDA"
    F_LIGHTORANG = "FCE4D6"
    F_LIGHTTEAL  = "D9EEF0"
    F_LIGHTYELL  = "FFF2CC"
    F_HOCH_BG    = "C6EFCE"
    T_HOCH       = "276221"
    F_MITTEL_BG  = "FFEB9C"
    T_MITTEL     = "9C5700"
    F_EXACT      = "D9E1F2"
    F_PHRASE     = "F4F4F4"

    # ---- Stil-Hilfsfunktionen ----
    def mk_fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def mk_font(size=10, bold=False, color=C_DARK):
        return Font(name="Arial", size=size, bold=bold, color=color)

    thin_side = Side(style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def mk_align(h="left", v="center", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def set_col_headers(ws, row, headers, fill=C_NAVY):
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = mk_font(10, True, C_WHITE)
            cell.fill = mk_fill(fill)
            cell.alignment = mk_align("center", "center", True)
            cell.border = thin_border
        ws.row_dimensions[row].height = 20

    def set_data_row(ws, row, values, fills, fonts=None, aligns=None):
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            if fills and ci <= len(fills) and fills[ci - 1]:
                cell.fill = mk_fill(fills[ci - 1])
            cell.font = (fonts[ci - 1] if fonts and ci <= len(fonts) and fonts[ci - 1] else mk_font())
            cell.alignment = (aligns[ci - 1] if aligns and ci <= len(aligns) and aligns[ci - 1] else mk_align())
            cell.border = thin_border
        ws.row_dimensions[row].height = 18

    # ---- Template-Header-Kopierfunktion ----
    def copy_template_header(tmpl_ws, dst_ws, num_header_rows, replace_map=None):
        import copy as _copy
        for col_letter, dim in tmpl_ws.column_dimensions.items():
            if dim.width:
                dst_ws.column_dimensions[col_letter].width = dim.width
        for merge in tmpl_ws.merged_cells.ranges:
            if merge.min_row <= num_header_rows:
                try:
                    dst_ws.merge_cells(start_row=merge.min_row, start_column=merge.min_col,
                                       end_row=min(merge.max_row, num_header_rows), end_column=merge.max_col)
                except Exception: pass
        for r in range(1, num_header_rows + 1):
            dst_ws.row_dimensions[r].height = tmpl_ws.row_dimensions[r].height or 15
            for c in range(1, tmpl_ws.max_column + 1):
                src = tmpl_ws.cell(row=r, column=c)
                dst = dst_ws.cell(row=r, column=c)
                # MergedCell ist nur ein Proxy ohne eigene Werte – ueberspringen
                if src.__class__.__name__ == "MergedCell":
                    continue
                val = src.value
                if val is not None and replace_map:
                    for old, new in replace_map.items():
                        if old in str(val): val = str(val).replace(old, new)
                dst.value = val
                if src.has_style:
                    dst.font = _copy.copy(src.font)
                    dst.fill = _copy.copy(src.fill)
                    dst.border = _copy.copy(src.border)
                    dst.alignment = _copy.copy(src.alignment)
        return num_header_rows + 1

    def fallback_banner(ws, row, text, font_size, bold, fill, merged_to, height):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merged_to)
        c = ws.cell(row=row, column=1, value=text)
        c.font = mk_font(font_size, bold, C_WHITE)
        c.fill = mk_fill(fill)
        c.alignment = mk_align("center", "center", False)
        ws.row_dimensions[row].height = height

    _tmpl_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "template", "adsforsocial_analyse.xlsx")
    _tmpl_wb = None
    try:
        _tmpl_wb = openpyxl.load_workbook(_tmpl_path)
    except Exception: pass

    domain = data.get("domain", "")
    datum = data.get("datum", datetime.datetime.now(ZoneInfo("Europe/Berlin")).strftime("%B %Y"))
    landingpage_analyse = data.get("landingpage_analyse", {})
    kampagnen = data.get("kampagnen", [])
    keywords = data.get("keywords", [])
    anzeigentexte = data.get("anzeigentexte", [])
    is_old_format = not (kampagnen or keywords or anzeigentexte or landingpage_analyse) and any(
        isinstance(v, list) for v in data.values()
    )

    wb = openpyxl.Workbook()
    if wb.sheetnames:
        wb.remove(wb.active)

    if is_old_format:
        # Rueckwaertskompatibilitaet: altes Zeilenformat
        for sheet_name, rows in data.items():
            clean_name = re.sub(r"[\\*?:/\[\]]", "", sheet_name)[:31] or "Tabelle"
            ws = wb.create_sheet(title=clean_name)
            ws.views.sheetView[0].showGridLines = False
            if not rows:
                continue
            for r_idx, row_data in enumerate(rows, start=1):
                ws.append(row_data)
                is_hdr = (r_idx == 1)
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=ci)
                    if is_hdr:
                        cell.font = mk_font(10, True, C_WHITE)
                        cell.fill = mk_fill(C_NAVY)
                        cell.alignment = mk_align("center", "center", True)
                    else:
                        cell.font = mk_font()
                        cell.alignment = mk_align()
                    cell.border = thin_border
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(min(ml + 4, 60), 12)

    else:
        # =================== Sheet 0: Landingpage-Analyse ===================
        if landingpage_analyse:
            ws = wb.create_sheet(title="Landingpage-Analyse")
            ws.views.sheetView[0].showGridLines = False

            _tmpl_lp = _tmpl_wb["Landingpage-Analyse"] if _tmpl_wb and "Landingpage-Analyse" in _tmpl_wb.sheetnames else None
            if _tmpl_lp:
                replace_map = {
                    "adsforsocial.de": domain if domain else "adsforsocial.de",
                    "Mai 2025": datum,
                }
                copy_template_header(_tmpl_lp, ws, _tmpl_lp.max_row, replace_map)

                # USPs befuellen (Zeilen 6-11)
                usps = landingpage_analyse.get("usps", [])
                for idx in range(6):
                    row = 6 + idx
                    if idx < len(usps):
                        usp = usps[idx]
                        ws.cell(row=row, column=1, value=str(idx + 1))
                        ws.cell(row=row, column=2, value=usp.get("bezeichnung", ""))
                        ws.cell(row=row, column=3, value=usp.get("beschreibung", ""))
                    else:
                        ws.cell(row=row, column=1, value="")
                        ws.cell(row=row, column=2, value="")
                        ws.cell(row=row, column=3, value="")

                # Zielgruppen befuellen (Zeilen 16-19)
                zielgruppen = landingpage_analyse.get("zielgruppen", [])
                for idx in range(4):
                    row = 16 + idx
                    if idx < len(zielgruppen):
                        zg = zielgruppen[idx]
                        ws.cell(row=row, column=1, value=str(idx + 1))
                        ws.cell(row=row, column=2, value=zg.get("name", ""))
                        ws.cell(row=row, column=3, value=zg.get("beschreibung", ""))
                    else:
                        ws.cell(row=row, column=1, value="")
                        ws.cell(row=row, column=2, value="")
                        ws.cell(row=row, column=3, value="")

                # Produkte befuellen (Zeilen 24-28)
                produkte = landingpage_analyse.get("produkte", [])
                for idx in range(5):
                    row = 24 + idx
                    if idx < len(produkte):
                        prod = produkte[idx]
                        ws.cell(row=row, column=1, value=prod.get("name", ""))
                        ws.cell(row=row, column=2, value=prod.get("inhalte", ""))
                        ws.cell(row=row, column=3, value=prod.get("formate", ""))
                        ws.cell(row=row, column=4, value=prod.get("besonderheiten", ""))
                    else:
                        ws.cell(row=row, column=1, value="")
                        ws.cell(row=row, column=2, value="")
                        ws.cell(row=row, column=3, value="")
                        ws.cell(row=row, column=4, value="")

                # Tonalitaet befuellen (Zeilen 33-35)
                tonalitaet = landingpage_analyse.get("tonalitaet", [])
                for idx in range(3):
                    row = 33 + idx
                    if idx < len(tonalitaet):
                        ton = tonalitaet[idx]
                        ws.cell(row=row, column=1, value=ton.get("merkmal", ""))
                        ws.cell(row=row, column=2, value=ton.get("auspraegung", ""))
                    else:
                        ws.cell(row=row, column=1, value="")
                        ws.cell(row=row, column=2, value="")

                # Keywords befuellen (Zeilen 39-43)
                lp_keywords = landingpage_analyse.get("keywords", [])
                kw_slots = [
                    (39, 1), (40, 1), (41, 1), (42, 1), (43, 1),
                    (39, 3), (40, 3), (41, 3), (42, 3),
                    (39, 5), (40, 5), (41, 5), (42, 5)
                ]
                for idx, (r, c) in enumerate(kw_slots):
                    if idx < len(lp_keywords):
                        ws.cell(row=r, column=c, value=lp_keywords[idx])
                    else:
                        ws.cell(row=r, column=c, value="")
            else:
                fallback_banner(ws, 1, f"Landingpage-Analyse – {domain}", 15, True, C_NAVY, 6, 24)
                ws.cell(row=3, column=1, value="Template-Blatt 'Landingpage-Analyse' nicht gefunden.")

        # =================== Sheet 1: Kampagnenstruktur ===================
        if kampagnen:
            NC = 5
            ws = wb.create_sheet(title="Kampagnenstruktur")
            ws.views.sheetView[0].showGridLines = False

            # Template-Blatt "Kampagnenstruktur" hat Zeilen 1-5 als Header
            # (1=Haupt-Banner, 2=Sub-Banner, 3=Leerzeile, 4=Sektions-Header, 5=Spalten-Header)
            _tmpl_kamp = _tmpl_wb["Kampagnenstruktur"] if _tmpl_wb and "Kampagnenstruktur" in _tmpl_wb.sheetnames else None
            if _tmpl_kamp:
                replace_map = {
                    "adsforsocial.de": domain if domain else "adsforsocial.de",
                    "Mai 2025": datum,
                }
                data_start_row = copy_template_header(_tmpl_kamp, ws, 5, replace_map)
                # Spalten-Header in Zeile 5 mit unseren Spalten ueberschreiben
                set_col_headers(ws, 5, ["Kampagne", "Kampagnentyp", "Budget", "Gebotsstrategie", "Begruendung"])
            else:
                # Fallback wenn Template fehlt
                for ci, w in enumerate([28, 22, 18, 28, 50], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                fallback_banner(ws, 1, f"Kampagnenstruktur-Empfehlung – {domain}", 15, True, C_NAVY, NC, 24)
                fallback_banner(ws, 2, f"Agent: strategy_agent  |  Plattform: Google Ads  |  Stand: {datum}", 9, False, C_MEDBLUE, NC, 16)
                ws.row_dimensions[3].height = 8
                set_col_headers(ws, 5, ["Kampagne", "Kampagnentyp", "Budget", "Gebotsstrategie", "Begruendung"])
                data_start_row = 6

            KAMP_FILLS = {
                "pmax": F_LIGHTGREEN, "performance max": F_LIGHTGREEN,
                "remarketing": F_LIGHTORANG, "retargeting": F_LIGHTORANG,
            }
            for ki, kamp in enumerate(kampagnen):
                row = data_start_row + ki
                name = kamp.get("name", "")
                fill = F_GRAY if ki % 2 == 0 else F_LIGHTBLUE
                for key, kf in KAMP_FILLS.items():
                    if key in name.lower():
                        fill = kf
                        break
                set_data_row(
                    ws, row,
                    [name, kamp.get("typ", ""), kamp.get("budget", ""),
                     kamp.get("gebotsstrategie", ""), kamp.get("begruendung", "")],
                    [fill] * NC,
                    fonts=[mk_font()] * NC,
                    aligns=[mk_align("left", "center", True)] * NC,
                )

        # =================== Sheet 2: Keyword-Recherche ===================
        if keywords:
            NC = 6
            ws = wb.create_sheet(title="Keyword-Recherche")
            ws.views.sheetView[0].showGridLines = False

            # Template-Blatt "Keyword-Recherche" hat Zeilen 1-4 als Header
            # (1=Haupt-Banner, 2=Sub-Banner, 3=Leerzeile, 4=Spalten-Header)
            _tmpl_kw = _tmpl_wb["Keyword-Recherche"] if _tmpl_wb and "Keyword-Recherche" in _tmpl_wb.sheetnames else None
            if _tmpl_kw:
                replace_map = {
                    "adsforsocial.de": domain if domain else "adsforsocial.de",
                    "Mai 2025": datum,
                }
                data_start_kw = copy_template_header(_tmpl_kw, ws, 4, replace_map)
            else:
                for ci, w in enumerate([28, 38, 12, 28, 22, 12], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                fallback_banner(ws, 1, f"Keyword-Recherche – {domain}", 15, True, C_NAVY, NC, 24)
                fallback_banner(ws, 2, f"Agent: keyword_agent  |  Plattform: Google Ads  |  Stand: {datum}", 9, False, C_MEDBLUE, NC, 16)
                ws.row_dimensions[3].height = 8
                set_col_headers(ws, 4, ["Kampagne / Anzeigengruppe", "Keyword", "Match Type", "Suchintention", "Kategorie", "Prioritaet"])
                data_start_kw = 5

            row = data_start_kw
            last_kamp_key = None
            KW_SEC_COLORS = [C_MEDBLUE, C_GREEN_DARK, C_ORANGE, C_TEAL, C_YELLOW_DK, C_NAVY]
            kamp_color_map = {}
            color_idx = 0

            for kw_idx, kw in enumerate(keywords):
                kampagne = kw.get("kampagne", "")
                ag = kw.get("anzeigengruppe", "")
                keyword = kw.get("keyword", "")
                mt = kw.get("match_type", "")
                sint = kw.get("suchintention", "")
                kat = kw.get("kategorie", "")
                prio = kw.get("prioritaet", "")
                kamp_key = f"{kampagne}|{ag}"

                if kamp_key != last_kamp_key:
                    if kampagne not in kamp_color_map:
                        kamp_color_map[kampagne] = KW_SEC_COLORS[color_idx % len(KW_SEC_COLORS)]
                        color_idx += 1
                    sec_color = kamp_color_map[kampagne]
                    sec_text = f"{kampagne} – {ag}" if ag else kampagne
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
                    c = ws.cell(row=row, column=1, value=sec_text)
                    c.font = mk_font(11, True, C_WHITE)
                    c.fill = mk_fill(sec_color)
                    c.alignment = mk_align("left", "center", False)
                    ws.row_dimensions[row].height = 18
                    row += 1
                    last_kamp_key = kamp_key

                base = F_GRAY if kw_idx % 2 == 0 else F_LIGHTBLUE
                mt_lower = mt.lower()
                if "exact" in mt_lower:
                    mt_fill, mt_font = F_EXACT, mk_font(10, True)
                elif "phrase" in mt_lower:
                    mt_fill, mt_font = F_PHRASE, mk_font(10, True)
                else:
                    mt_fill, mt_font = base, mk_font()

                if "hoch" in prio.lower():
                    pf, ptxt = F_HOCH_BG, mk_font(10, True, T_HOCH)
                elif "mittel" in prio.lower():
                    pf, ptxt = F_MITTEL_BG, mk_font(10, True, T_MITTEL)
                else:
                    pf, ptxt = base, mk_font()

                set_data_row(
                    ws, row,
                    [kampagne, keyword, mt, sint, kat, prio],
                    [base, base, mt_fill, base, base, pf],
                    fonts=[mk_font(), mk_font(), mt_font, mk_font(), mk_font(), ptxt],
                    aligns=[
                        mk_align(), mk_align(),
                        mk_align("center", "center", False),
                        mk_align(), mk_align(),
                        mk_align("center", "center", False),
                    ],
                )
                row += 1

        # =================== Sheet 3: Anzeigentexte (RSA) ===================
        if anzeigentexte:
            NC = 6
            ws = wb.create_sheet(title="Anzeigentexte (RSA)")
            ws.views.sheetView[0].showGridLines = False

            # Template-Blatt "Anzeigentexte (RSA)" hat Zeilen 1-2 als Header
            # (1=Haupt-Banner, 2=Sub-Banner)
            _tmpl_at = _tmpl_wb["Anzeigentexte (RSA)"] if _tmpl_wb and "Anzeigentexte (RSA)" in _tmpl_wb.sheetnames else None
            if _tmpl_at:
                replace_map = {
                    "adsforsocial.de": domain if domain else "adsforsocial.de",
                    "Mai 2025": datum,
                }
                data_start_at = copy_template_header(_tmpl_at, ws, 2, replace_map)
            else:
                for ci, w in enumerate([22, 50, 10, 8, 12, 20], 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w
                fallback_banner(ws, 1, f"Anzeigentexte (RSA) – {domain}", 15, True, C_NAVY, NC, 24)
                fallback_banner(ws, 2, f"Agent: copywriter_agent  |  Plattform: Google Ads  |  Stand: {datum}", 9, False, C_MEDBLUE, NC, 16)
                data_start_at = 3

            row = data_start_at
            AG_SEC_COLORS = [C_MEDBLUE, C_GREEN_DARK, C_ORANGE, C_TEAL, C_YELLOW_DK, C_NAVY]
            AG_ROW_FILLS  = [F_LIGHTBLUE, F_LIGHTGREEN, F_LIGHTORANG, F_LIGHTTEAL, F_LIGHTYELL, F_GRAY]
            POS_LABELS = [
                "Pos. 1 – Produkt / Keyword",
                "Pos. 2 – USP / Nutzen",
                "Pos. 3 – Call-to-Action",
            ]

            for ag_idx, ag in enumerate(anzeigentexte):
                ag_name     = ag.get("anzeigengruppe", f"Anzeigengruppe {ag_idx + 1}")
                kampagne    = ag.get("kampagne", "")
                headlines   = ag.get("headlines", [])
                descriptions = ag.get("descriptions", [])
                sec_color   = AG_SEC_COLORS[ag_idx % len(AG_SEC_COLORS)]
                row_fill    = AG_ROW_FILLS[ag_idx % len(AG_ROW_FILLS)]

                # Leerzeile + Abschnitts-Header
                ws.row_dimensions[row].height = 8
                row += 1
                sec_label = f"{kampagne} – {ag_name}" if kampagne else ag_name
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
                c = ws.cell(row=row, column=1, value=sec_label)
                c.font = mk_font(11, True, C_WHITE)
                c.fill = mk_fill(sec_color)
                c.alignment = mk_align("left", "center", False)
                ws.row_dimensions[row].height = 18
                row += 1

                if headlines:
                    # Headlines Sub-Header
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
                    c = ws.cell(row=row, column=1, value="Headlines (max. 30 Zeichen)")
                    c.font = mk_font(10, True, C_WHITE)
                    c.fill = mk_fill(sec_color)
                    c.alignment = mk_align("left", "center", False)
                    ws.row_dimensions[row].height = 16
                    row += 1

                    set_col_headers(ws, row, ["Position", "Headline-Text", "Zeichen", "Limit", "Status", "Notizen"])
                    row += 1

                    for hl_idx, hl in enumerate(headlines):
                        pos_label = POS_LABELS[hl_idx // 5] if hl_idx // 5 < len(POS_LABELS) and hl_idx % 5 == 0 else ""
                        ch = len(hl)
                        ok = ch <= 30
                        zf = "E2EFDA" if ok else "FCE4D6"
                        sf = "E2EFDA" if ok else "FCE4D6"
                        zt = mk_font(9, True, C_GREEN_DARK if ok else C_ORANGE)
                        st = mk_font(10, True, C_GREEN_DARK if ok else C_ORANGE)
                        rf = row_fill if hl_idx % 2 == 0 else F_GRAY
                        set_data_row(
                            ws, row,
                            [pos_label, hl, f"{ch}/30", "30", "OK" if ok else "LIMIT", ""],
                            [rf, rf, zf, rf, sf, rf],
                            fonts=[mk_font(10, bool(pos_label)), mk_font(), zt, mk_font(9), st, mk_font()],
                            aligns=[
                                mk_align("left", "center", False),
                                mk_align("left", "center", False),
                                mk_align("center", "center", False),
                                mk_align("center", "center", False),
                                mk_align("center", "center", False),
                                mk_align(),
                            ],
                        )
                        row += 1

                if descriptions:
                    ws.row_dimensions[row].height = 8
                    row += 1
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
                    c = ws.cell(row=row, column=1, value="Descriptions (max. 90 Zeichen)")
                    c.font = mk_font(10, True, C_WHITE)
                    c.fill = mk_fill(sec_color)
                    c.alignment = mk_align("left", "center", False)
                    ws.row_dimensions[row].height = 16
                    row += 1

                    set_col_headers(ws, row, ["Nr.", "Description-Text", "Zeichen", "Limit", "Status", "Notizen"])
                    row += 1

                    for di, desc in enumerate(descriptions):
                        ch = len(desc)
                        ok = ch <= 90
                        zf = "E2EFDA" if ok else "FCE4D6"
                        sf = "E2EFDA" if ok else "FCE4D6"
                        zt = mk_font(9, True, C_GREEN_DARK if ok else C_ORANGE)
                        st = mk_font(10, True, C_GREEN_DARK if ok else C_ORANGE)
                        rf = row_fill if di % 2 == 0 else F_GRAY
                        set_data_row(
                            ws, row,
                            [f"Desc. {di + 1}", desc, f"{ch}/90", "90", "OK" if ok else "LIMIT", ""],
                            [rf, rf, zf, rf, sf, rf],
                            fonts=[mk_font(10, True), mk_font(), zt, mk_font(9), st, mk_font()],
                            aligns=[
                                mk_align("center", "center", False),
                                mk_align("left", "center", True),
                                mk_align("center", "center", False),
                                mk_align("center", "center", False),
                                mk_align("center", "center", False),
                                mk_align(),
                            ],
                        )
                        row += 1

    wb.save(filepath)
    return f"Kampagnendaten erfolgreich als Excel (.xlsx) exportiert unter {filepath}."


def create_excel_exporter_agent():
    return Agent(
        name="excel_exporter_agent",
        model=Gemini(
            model="gemini-flash-latest",
            retry_options=types.HttpRetryOptions(attempts=3),
        ),
        instruction=apply_rules("""Du bist der Excel-Export-Spezialist. Deine Aufgabe ist es, die vollstaendige Google Ads Kampagnenstruktur sowie die Landingpage-Analyse in ein semantisches JSON-Format umzuwandeln und als professionell gestaltete Excel-Datei zu speichern.

Erstelle ein JSON-Objekt mit folgender Struktur und rufe damit das `save_campaign_as_excel`-Tool auf:

{
  "domain": "<domain des Kunden, z.B. adsforsocial.de>",
  "datum": "<aktueller Monat und Jahr, z.B. Mai 2025>",
  "landingpage_analyse": {
    "usps": [
      {
        "bezeichnung": "<Kurze USP Bezeichnung, z.B. Schnelligkeit>",
        "beschreibung": "<Beschreibung des USPs>"
      }
    ],
    "zielgruppen": [
      {
        "name": "<Zielgruppen-Name, z.B. E-Commerce & Onlineshops>",
        "beschreibung": "<Beschreibung der Zielgruppe>"
      }
    ],
    "produkte": [
      {
        "name": "<Paket / Leistung, z.B. Basic Paket>",
        "inhalte": "<Inhalt/Menge/Details>",
        "formate": "<Formate, z.B. FB, IG, Story>",
        "besonderheiten": "<Besonderheiten/USPs des Produkts>"
      }
    ],
    "tonalitaet": [
      {
        "merkmal": "<Merkmal, z.B. Kommunikationsstil, Ansprache, Ausrichtung>",
        "auspraegung": "<Beschreibung der Auspraegung>"
      }
    ],
    "keywords": [
      "<Keyword 1>",
      "<Keyword 2>"
    ]
  },
  "kampagnen": [
    {
      "name": "<Kampagnenname, z.B. Brand Search>",
      "typ": "<Kampagnentyp, z.B. Suchnetzwerk>",
      "budget": "<Budget, z.B. 15 Euro/Tag>",
      "gebotsstrategie": "<Gebotsstrategie, z.B. Klicks maximieren>",
      "begruendung": "<Strategische Begruendung in 1-2 Saetzen>"
    }
  ],
  "keywords": [
    {
      "kampagne": "<Kampagnenname>",
      "anzeigengruppe": "<Anzeigengruppenname>",
      "keyword": "<Keyword inklusive Match-Type-Klammern, z.B. [adsforsocial] oder 'ads for social'>",
      "match_type": "<Exact, Phrase oder Broad>",
      "suchintention": "<Suchintention, z.B. Navigational / Brand>",
      "kategorie": "<Kategorie, z.B. Brand>",
      "prioritaet": "<Hoch, Mittel oder Niedrig>"
    }
  ],
  "anzeigentexte": [
    {
      "kampagne": "<Kampagnenname>",
      "anzeigengruppe": "<Anzeigengruppenname>",
      "headlines": ["<Headline 1 max. 30 Zeichen>", "<Headline 2 max. 30 Zeichen>"],
      "descriptions": ["<Description 1 max. 90 Zeichen>", "<Description 2 max. 90 Zeichen>"]
    }
  ]
}

WICHTIG:
- Extrahiere die USPs, Zielgruppen, angebotenen Produkte, Marken-Tonalitaet und Keywords aus der Landingpage-Analyse des landing_page_agent und befulle 'landingpage_analyse' praezise gemaess dem Schema. Limitiere dich auf max. 6 USPs, 4 Zielgruppen, 5 Produkte, 3 Tonalitaetsmerkmale und 13 Keywords, damit das feste Tabellenlayout im Template optimal genutzt wird.
- Fasse alle Kampagnen aus strategy_agent unter 'kampagnen' zusammen (eine Zeile pro Kampagne).
- Liste alle Keywords aus keyword_agent unter 'keywords' auf.
- Erstelle fuer jede Anzeigengruppe des copywriter_agent einen eigenen Eintrag unter 'anzeigentexte'.
- Headlines MUESSEN unter 30 Zeichen bleiben. Descriptions MUESSEN unter 90 Zeichen bleiben.
- Waehle einen sinnvollen Dateinamen basierend auf der Domain (z.B. kampagne_adsforsocial.xlsx).
- Gib nach dem Export nur den absoluten Pfad der exportierten Datei zurueck."""),
        description="Konvertiert Kampagnen-Strategien, Keywords, Anzeigentexte und die Landingpage-Analyse in eine formatierte Excel-Datei (.xlsx) mit mehreren Tabellenblaettern.",
        tools=[save_campaign_as_excel],
    )
